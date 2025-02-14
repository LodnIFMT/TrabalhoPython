import customtkinter as ctk
import openpyxl
import json

def procurar_jogo(app): #Função para prucurar o Jogo
    try: #Tenta executar o código
        with open("data.json", "r") as f: #Abre o arquivo "data.json"
            dados = json.load(f) #A Variavel dados recebe o dicionario do arquivo "data.json"
            planilha = dados["Planilha"] #Recebe o caminho da planilha
            jogo = dados["Jogo"] #Recebe informações do jogo

        lista_jogos = [] #Lista para receber a lista dos jogos filtrados

        wb = openpyxl.load_workbook(planilha) #Abre o Arquivo Excel
        sheet= wb.active #Variavel que representa a planilha

        for linha in sheet.iter_rows(min_col=3, min_row=2, max_col=8, max_row=sheet.max_row): #Para cada linha na planilha faça:
            if (jogo["Nome"] in str(linha[0].value).lower() if jogo["Nome"] else True) and (jogo["Desenvolvedor"].lower() in str(linha[5].value).lower() if jogo["Desenvolvedor"] else True) and (jogo["Plataforma"] in str(linha[1].value) if jogo["Plataforma"] else True) and (str(jogo["Ano"]) in str(linha[2].value) if jogo["Ano"] else True):
                lista_jogos.append(str(linha[0].value)) #Adiciona o nome do jogo na lista jogos

        jogos_set = sorted(set(lista_jogos))

        jogos = []
        for i in jogos_set:
            jogos.append(i)
                
        app.lista_jogo.delete(0, ctk.END)

        for jg in jogos:
            itens = app.lista_jogo.size()
            
            if itens < 20:
                app.lista_jogo.insert(ctk.END, jg)
            else:
                break
        
        return jogos

    except FileNotFoundError: #caso não achar o Arquivo "data.json" faça os print a segui
        print('Arquivo "data.json" não encontrado ...')
        print('Planilha não encontrada ...')
        print('Criando arquivo "data.json ..."')

    
def Desenvolvedor():
    try:
        with open("data.json", "r") as f:
            dados = json.load(f)
            planilha = dados["Planilha"]
        
        wb = openpyxl.load_workbook(planilha)
        sheet = wb.active

        devs_planilha = []
        for linha in sheet.iter_rows(min_col=8, min_row=2, max_row=sheet.max_row):
            devs_planilha.extend(str(linha[0].value).split("/"))

        devs_set = sorted(set(devs_planilha))

        devs = []
        for i in devs_set:
            devs.append(i)

        return devs
    
    except FileNotFoundError:
        print('Arquivo "data.json" não encontrado ...')

def bt_devs(jogo:str):
    with open("data.json", "r") as f:
        arquivo = json.load(f)
        planilha = arquivo["Planilha"]

    wb = openpyxl.load_workbook(planilha)
    sheet = wb.active

    devs = []
    for linha in sheet.iter_rows(min_col=3, max_col=8, min_row=2, max_row=sheet.max_row):
        if linha[0].value == jogo:
            devs.append(str(linha[5].value).split('/'))

    return devs


def plataforma():
    try:
        with open("data.json", "r") as f:
            dados = json.load(f)
            planilha = dados["Planilha"]
        
        wb = openpyxl.load_workbook(planilha)
        sheet = wb.active

        pltf_planilha = []
        for linha in sheet.iter_rows(min_col=4, min_row=2, max_row=sheet.max_row):
            pltf_planilha.append(linha[0].value)

        pltf_set = sorted(set(pltf_planilha))

        pltf = []
        for i in pltf_set:
            pltf.append(i)

        return pltf
    
    except FileNotFoundError:
        print('Arquivo "data.json" não encontrado ...')

def ano():
    try:
        with open("data.json", "r") as f:
            dados = json.load(f)
            planilha = dados["Planilha"]

        wb = openpyxl.load_workbook(planilha)
        sheet = wb.active
        
        anos_planilha = []
        for linha in sheet.iter_rows(min_col=5, min_row=2, max_row=sheet.max_row):
            anos_planilha.append(str(linha[0].value))
        
        anos_set = sorted(set(anos_planilha))
        
        anos = []
        for i in anos_set:
            anos.append(i)

        return anos

    except FileNotFoundError:
        print('Arquivo "data.json" não encontrado ...')

def fun_proximo(app):
    app.lista_jogo.delete(0, ctk.END)
    app.indice = app.indice +20
        
    for i in range(app.indice, app.indice + 20):
        if i < len(app.lista)-1:
            app.lista_jogo.insert(ctk.END, app.lista[i])
        else:
            if app.indice < len(app.lista):
                app.lista_jogo.insert(ctk.END, app.lista[i])
                app.indice = 0
    
def fun_voltar(app):
    app.lista_jogo.delete(0, ctk.END)
    app.indice = app.indice -20
        
    if app.indice < 0:
        app.indice = 0

    for i in range(app.indice, app.indice + 20):
        app.lista_jogo.insert(ctk.END, app.lista[i])

def fun_limpar(app):
    app.info["Jogo"]["Nome"] = None
    app.info["Jogo"]["Ano"] = None
    app.info["Jogo"]["Desenvolvedor"] = None
    app.info["Jogo"]["Plataforma"] = None

    with open("data.json", "w") as f:
        json.dump(app.info, f, indent=2)

    app.nome_jogo.delete(0, ctk.END)
    app.opcao_ano.set("-----")
    app.opcao_devs.set("-----")
    app.opcao_plataforma.set("-----")
    app.lista_jogo.delete(0, ctk.END)

def informacoes(jogo:str):
    with open("data.json", "r") as f:
        arquivo = json.load(f)
        planilha = arquivo["Planilha"]

    wb = openpyxl.load_workbook(planilha)
    sheet = wb.active

    info = {
        "Nome": jogo,
        "Plataforma":[],
        "Desenvolvedor":[],
        "Ano":None
    }

    ano =[]
    for linha in sheet.iter_rows(min_col=3, max_col=8, min_row=2, max_row=sheet.max_row):
        if linha[0].value == jogo:
            info["Plataforma"].append(linha[1].value)
            info["Desenvolvedor"].append(str(linha[5].value))
            ano.append(linha[2].value)
    
    ano.sort()
    info["Ano"] = str(ano[0])

    arquivo["jg_info"] = info
    with open("data.json", "w") as f:
        json.dump(arquivo, f, indent=2)